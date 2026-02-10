"""Generate Excel reports from analysis data."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .report_data import ReportDataBuilder, ServerReport

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel reports with multiple sheets and charts."""

    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    OVERSIZED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    UNDERSIZED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    CURRENCY_FORMAT = "$#,##0.00"
    PERCENT_FORMAT = "0.0%"
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, output_path: Union[str, Path]):
        """Initialize the generator.

        Args:
            output_path: Path for the output Excel file
        """
        self.output_path = Path(output_path)
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)

    def generate(self, report_builder: ReportDataBuilder) -> Path:
        """Generate the complete Excel report.

        Args:
            report_builder: ReportDataBuilder with all data

        Returns:
            Path to generated file
        """
        data = report_builder.get_all_data()

        # Create sheets
        self._create_summary_sheet(data["summary"])
        self._create_server_details_sheet(data["servers"])
        self._create_recommendations_sheet(data["servers"])
        self._create_cost_analysis_sheet(data)
        self._create_contention_sheet(data["contention"])

        # Save workbook
        self.workbook.save(self.output_path)
        logger.info(f"Excel report saved to {self.output_path}")
        return self.output_path

    def _create_summary_sheet(self, summary: Dict[str, Any]) -> None:
        """Create the Executive Summary sheet."""
        ws = self.workbook.create_sheet("Executive Summary")

        # Title
        ws["A1"] = "AWS Cost Optimization Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        ws["A2"] = f"Generated: {summary['generated_at']}"
        ws["A2"].font = Font(italic=True)

        # Key Metrics
        row = 4
        metrics = [
            ("Total Servers Analyzed", summary["total_servers"], None),
            ("Current Monthly Spend", summary["total_current_monthly"], self.CURRENCY_FORMAT),
            ("Potential Monthly Savings", summary["total_monthly_savings"], self.CURRENCY_FORMAT),
            ("Potential Yearly Savings", summary["total_yearly_savings"], self.CURRENCY_FORMAT),
            ("Savings Percentage", summary["savings_percentage"] / 100, self.PERCENT_FORMAT),
        ]

        for label, value, fmt in metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=2, value=value)
            if fmt:
                cell.number_format = fmt
            row += 1

        # Classification Breakdown
        row += 1
        ws.cell(row=row, column=1, value="Server Classification").font = Font(bold=True, size=12)
        row += 1

        classifications = [
            ("Oversized (downsize candidates)", summary["oversized_count"], self.OVERSIZED_FILL),
            ("Right-sized (no change)", summary["right_sized_count"], None),
            ("Undersized (upsize candidates)", summary["undersized_count"], self.UNDERSIZED_FILL),
            ("With Resource Contention", summary["contention_count"], None),
        ]

        for label, value, fill in classifications:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            if fill:
                cell.fill = fill
            row += 1

        # Top 10 Savings Opportunities
        row += 1
        ws.cell(row=row, column=1, value="Top 10 Savings Opportunities").font = Font(bold=True, size=12)
        row += 1

        headers = ["Server", "Current Type", "Recommended", "Monthly Savings"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        row += 1

        for server in summary.get("top_10_savings", []):
            ws.cell(row=row, column=1, value=server.get("hostname") or server.get("server_id"))
            ws.cell(row=row, column=2, value=server.get("instance_type"))
            ws.cell(row=row, column=3, value=server.get("recommended_type"))
            cell = ws.cell(row=row, column=4, value=server.get("monthly_savings"))
            cell.number_format = self.CURRENCY_FORMAT
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 18

    def _create_server_details_sheet(self, servers: List[Dict]) -> None:
        """Create the Server Details sheet."""
        ws = self.workbook.create_sheet("Server Details")

        # Define columns
        columns = [
            ("Server", "hostname", 25),
            ("Instance ID", "instance_id", 20),
            ("Instance Type", "instance_type", 15),
            ("vCPU", "vcpu", 8),
            ("RAM (GB)", "memory_gb", 10),
            ("CPU Avg %", "cpu_avg", 12),
            ("CPU P95 %", "cpu_p95", 12),
            ("Mem Avg %", "memory_avg", 12),
            ("Mem P95 %", "memory_p95", 12),
            ("Disk Avg %", "disk_avg", 12),
            ("Data Days", "data_days", 10),
            ("Classification", "classification", 15),
            ("Recommended Type", "recommended_type", 15),
            ("Current Monthly", "current_monthly", 15),
            ("Monthly Savings", "monthly_savings", 15),
            ("Confidence", "confidence", 12),
            ("Risk Level", "risk_level", 10),
            ("GSI", "GSI", 15),
            ("Environment", "Environment", 15),
        ]

        # Headers
        for col, (header, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            ws.column_dimensions[cell.column_letter].width = width

        # Data rows
        for row_num, server in enumerate(servers, 2):
            for col, (header, field, _) in enumerate(columns, 1):
                # Check if field is a direct key or nested in tags
                if field in ["GSI", "Environment", "Team"]:
                    # These might be at top level (from report_data.to_dict()) or in tags
                    value = server.get(field) or server.get("tags", {}).get(field, "")
                else:
                    value = server.get(field)

                cell = ws.cell(row=row_num, column=col, value=value)

                # Apply conditional formatting for classification
                if field == "classification":
                    if value == "oversized":
                        cell.fill = self.OVERSIZED_FILL
                    elif value == "undersized":
                        cell.fill = self.UNDERSIZED_FILL

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_recommendations_sheet(self, servers: List[Dict]) -> None:
        """Create the Recommendations sheet."""
        ws = self.workbook.create_sheet("Recommendations")

        columns = [
            ("Server", "hostname", 25),
            ("Current Type", "instance_type", 15),
            ("Recommended", "recommended_type", 15),
            ("Classification", "classification", 15),
            ("Risk", "risk_level", 10),
            ("Confidence", "confidence", 12),
            ("Current Monthly", "current_monthly", 15),
            ("Recommended Monthly", "recommended_monthly", 18),
            ("Monthly Savings", "monthly_savings", 15),
            ("Yearly Savings", "yearly_savings", 15),
            ("Reason", "reason", 50),
        ]

        # Headers
        for col, (header, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            ws.column_dimensions[cell.column_letter].width = width

        # Filter to servers with recommendations
        with_recs = [s for s in servers if s.get("recommended_type")]

        # Sort by savings
        with_recs.sort(key=lambda x: x.get("monthly_savings", 0), reverse=True)

        # Data rows
        for row_num, server in enumerate(with_recs, 2):
            for col, (_, field, _) in enumerate(columns, 1):
                value = server.get(field)
                cell = ws.cell(row=row_num, column=col, value=value)

                # Format currency columns
                if field in ["current_monthly", "recommended_monthly", "monthly_savings", "yearly_savings"]:
                    cell.number_format = self.CURRENCY_FORMAT

                # Format confidence as percentage
                if field == "confidence":
                    cell.number_format = self.PERCENT_FORMAT

                # Apply conditional formatting
                if field == "classification":
                    if value == "oversized":
                        cell.fill = self.OVERSIZED_FILL
                    elif value == "undersized":
                        cell.fill = self.UNDERSIZED_FILL

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_cost_analysis_sheet(self, data: Dict[str, Any]) -> None:
        """Create the Cost Analysis sheet."""
        ws = self.workbook.create_sheet("Cost Analysis")

        summary = data["summary"]
        by_gsi = data.get("by_gsi", {})

        # Overall metrics
        ws["A1"] = "Cost Analysis Overview"
        ws["A1"].font = Font(size=14, bold=True)

        row = 3
        metrics = [
            ("Total Current Monthly Spend", summary["total_current_monthly"]),
            ("Total After Optimization", summary["total_current_monthly"] - summary["total_monthly_savings"]),
            ("Monthly Savings", summary["total_monthly_savings"]),
            ("Yearly Savings", summary["total_yearly_savings"]),
        ]

        for label, value in metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = self.CURRENCY_FORMAT
            row += 1

        # Savings by GSI
        row += 2
        ws.cell(row=row, column=1, value="Savings by GSI/Cost Center").font = Font(bold=True, size=12)
        row += 1

        headers = ["GSI", "Server Count", "Current Monthly", "Potential Savings"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        row += 1

        chart_start_row = row

        for gsi, values in by_gsi.items():
            ws.cell(row=row, column=1, value=gsi)
            ws.cell(row=row, column=2, value=values["count"])
            cell = ws.cell(row=row, column=3, value=values["current_monthly"])
            cell.number_format = self.CURRENCY_FORMAT
            cell = ws.cell(row=row, column=4, value=values["monthly_savings"])
            cell.number_format = self.CURRENCY_FORMAT
            row += 1

        chart_end_row = row - 1

        # Add bar chart for savings by GSI
        if chart_end_row >= chart_start_row:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Monthly Savings by GSI"
            chart.y_axis.title = "Savings ($)"

            data_ref = Reference(ws, min_col=4, min_row=chart_start_row - 1,
                                max_row=chart_end_row, max_col=4)
            cats = Reference(ws, min_col=1, min_row=chart_start_row,
                           max_row=chart_end_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4

            ws.add_chart(chart, "F3")

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

    def _create_contention_sheet(self, contention_data: Dict[str, Any]) -> None:
        """Create the Contention Report sheet."""
        ws = self.workbook.create_sheet("Contention Report")

        # Summary
        ws["A1"] = "Resource Contention Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        row = 3
        ws.cell(row=row, column=1, value="Servers with Contention").font = Font(bold=True)
        ws.cell(row=row, column=2, value=contention_data["total_with_contention"])
        row += 1

        ws.cell(row=row, column=1, value="Total Contention Events").font = Font(bold=True)
        ws.cell(row=row, column=2, value=contention_data["total_contention_events"])
        row += 1

        ws.cell(row=row, column=1, value="Total Contention Hours").font = Font(bold=True)
        ws.cell(row=row, column=2, value=round(contention_data["total_contention_hours"], 1))
        row += 2

        # Servers with contention
        if contention_data["servers"]:
            ws.cell(row=row, column=1, value="Servers with Resource Contention").font = Font(bold=True, size=12)
            row += 1

            headers = ["Server", "Instance Type", "Events", "Contention Hours",
                      "CPU P95 %", "Mem P95 %", "Classification"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            row += 1

            for server in contention_data["servers"]:
                ws.cell(row=row, column=1, value=server.get("hostname") or server.get("server_id"))
                ws.cell(row=row, column=2, value=server.get("instance_type"))
                ws.cell(row=row, column=3, value=server.get("contention_events"))
                ws.cell(row=row, column=4, value=server.get("contention_hours"))
                ws.cell(row=row, column=5, value=server.get("cpu_p95"))
                ws.cell(row=row, column=6, value=server.get("memory_p95"))
                ws.cell(row=row, column=7, value=server.get("classification"))
                row += 1
        else:
            ws.cell(row=row, column=1, value="No servers with resource contention detected.")

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 15


def generate_excel_report(
    report_builder: ReportDataBuilder,
    output_path: Union[str, Path]
) -> Path:
    """Convenience function to generate an Excel report.

    Args:
        report_builder: ReportDataBuilder with all data
        output_path: Path for the output file

    Returns:
        Path to generated file
    """
    generator = ExcelGenerator(output_path)
    return generator.generate(report_builder)

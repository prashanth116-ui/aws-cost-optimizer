"""Multi-account report generation."""

import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from .report_data import ServerReport, ReportDataBuilder
from .excel_generator import ExcelGenerator
from ..clients.multi_account_client import MultiAccountSummary, AccountAnalysisResult

logger = logging.getLogger(__name__)


@dataclass
class MultiAccountServerReport(ServerReport):
    """Server report with multi-account context."""

    account_id: str = ""
    account_name: str = ""

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary with account info."""
        base = super().to_dict()
        base["account_id"] = self.account_id
        base["account_name"] = self.account_name
        return base


class MultiAccountReportBuilder(ReportDataBuilder):
    """Build reports from multi-account analysis.

    Extends ReportDataBuilder to handle cross-account data
    with account-level aggregation.
    """

    def __init__(self):
        """Initialize the builder."""
        super().__init__()
        self.account_summaries: Dict[str, Dict[str, Any]] = {}

    def add_account_data(
        self,
        result: AccountAnalysisResult,
        server_reports: List[ServerReport]
    ) -> None:
        """Add data from an account analysis.

        Args:
            result: AccountAnalysisResult from multi-account client
            server_reports: List of ServerReport objects for the account
        """
        # Add servers with account context
        for report in server_reports:
            enhanced_report = MultiAccountServerReport(
                **{**report.__dict__},
                account_id=result.account_id,
                account_name=result.account_name,
            )
            self.add_server(enhanced_report)

        # Store account summary
        self.account_summaries[result.account_id] = {
            "account_id": result.account_id,
            "account_name": result.account_name,
            "instance_count": len(result.instances),
            "monthly_cost": result.costs.get("total_monthly", 0),
            "success": result.success,
            "error": result.error,
        }

    def build_by_account(self) -> Dict[str, Dict[str, Any]]:
        """Build summary grouped by account.

        Returns:
            Dictionary mapping account_id to summary data
        """
        by_account: Dict[str, Dict[str, Any]] = {}

        for server in self.servers:
            account_id = getattr(server, "account_id", "unknown")
            account_name = getattr(server, "account_name", "Unknown")

            if account_id not in by_account:
                by_account[account_id] = {
                    "account_id": account_id,
                    "account_name": account_name,
                    "count": 0,
                    "current_monthly": 0.0,
                    "monthly_savings": 0.0,
                    "oversized": 0,
                    "right_sized": 0,
                    "undersized": 0,
                    "with_contention": 0,
                }

            by_account[account_id]["count"] += 1
            by_account[account_id]["current_monthly"] += server.current_monthly
            by_account[account_id]["monthly_savings"] += server.monthly_savings

            if server.classification == "oversized":
                by_account[account_id]["oversized"] += 1
            elif server.classification == "right_sized":
                by_account[account_id]["right_sized"] += 1
            elif server.classification == "undersized":
                by_account[account_id]["undersized"] += 1

            if server.has_contention:
                by_account[account_id]["with_contention"] += 1

        # Sort by savings (descending)
        return dict(sorted(
            by_account.items(),
            key=lambda x: x[1]["monthly_savings"],
            reverse=True
        ))

    def build_summary(self) -> Dict[str, Any]:
        """Build summary including multi-account data.

        Returns:
            Summary dictionary with account breakdown
        """
        base_summary = super().build_summary()

        by_account = self.build_by_account()
        base_summary["by_account"] = by_account
        base_summary["account_count"] = len(by_account)

        # Top 5 accounts by savings
        top_accounts = list(by_account.values())[:5]
        base_summary["top_accounts_by_savings"] = [
            {
                "account_id": a["account_id"],
                "account_name": a["account_name"],
                "monthly_savings": round(a["monthly_savings"], 2),
                "instance_count": a["count"],
            }
            for a in top_accounts
        ]

        return base_summary


class MultiAccountExcelGenerator(ExcelGenerator):
    """Excel generator for multi-account reports.

    Extends ExcelGenerator to add account-specific sheets
    and cross-account analysis.
    """

    ACCOUNT_HEADER_FILL = PatternFill(
        start_color="2E7D32",
        end_color="2E7D32",
        fill_type="solid"
    )

    def generate(self, report_builder: MultiAccountReportBuilder) -> Path:
        """Generate the multi-account Excel report.

        Args:
            report_builder: MultiAccountReportBuilder with all data

        Returns:
            Path to generated file
        """
        data = report_builder.get_all_data()

        # Create standard sheets
        self._create_summary_sheet(data["summary"])
        self._create_account_summary_sheet(data["summary"])
        self._create_server_details_sheet(data["servers"])
        self._create_recommendations_sheet(data["servers"])
        self._create_cost_analysis_sheet(data)
        self._create_contention_sheet(data["contention"])

        # Save workbook
        self.workbook.save(self.output_path)
        logger.info(f"Multi-account report saved to {self.output_path}")
        return self.output_path

    def _create_account_summary_sheet(self, summary: Dict[str, Any]) -> None:
        """Create the Account Summary sheet."""
        ws = self.workbook.create_sheet("Account Summary")

        # Title
        ws["A1"] = "Multi-Account Summary"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:E1")

        ws["A2"] = f"Accounts Analyzed: {summary.get('account_count', 0)}"
        ws["A2"].font = Font(italic=True)

        # Account table
        row = 4
        headers = [
            "Account ID", "Account Name", "Instances",
            "Current Monthly", "Potential Savings", "Savings %"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.ACCOUNT_HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
        row += 1

        by_account = summary.get("by_account", {})

        for account_id, data in by_account.items():
            ws.cell(row=row, column=1, value=account_id)
            ws.cell(row=row, column=2, value=data["account_name"])
            ws.cell(row=row, column=3, value=data["count"])

            cell = ws.cell(row=row, column=4, value=data["current_monthly"])
            cell.number_format = self.CURRENCY_FORMAT

            cell = ws.cell(row=row, column=5, value=data["monthly_savings"])
            cell.number_format = self.CURRENCY_FORMAT

            savings_pct = (
                data["monthly_savings"] / data["current_monthly"] * 100
                if data["current_monthly"] > 0 else 0
            )
            cell = ws.cell(row=row, column=6, value=savings_pct / 100)
            cell.number_format = self.PERCENT_FORMAT

            row += 1

        # Add pie chart for savings distribution
        chart_start = 4
        chart_end = row - 1

        if chart_end > chart_start:
            chart = PieChart()
            chart.title = "Savings by Account"

            data_ref = Reference(ws, min_col=5, min_row=chart_start,
                                max_row=chart_end, max_col=5)
            labels = Reference(ws, min_col=2, min_row=chart_start + 1,
                              max_row=chart_end)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels)

            ws.add_chart(chart, "H3")

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 12

    def _create_server_details_sheet(self, servers: List[Dict]) -> None:
        """Create the Server Details sheet with account columns."""
        ws = self.workbook.create_sheet("Server Details")

        # Define columns (add account columns at start)
        columns = [
            ("Account", "account_name", 20),
            ("Server", "hostname", 25),
            ("Instance ID", "instance_id", 20),
            ("Instance Type", "instance_type", 15),
            ("vCPU", "vcpu", 8),
            ("RAM (GB)", "memory_gb", 10),
            ("CPU P95 %", "cpu_p95", 12),
            ("Mem P95 %", "memory_p95", 12),
            ("Classification", "classification", 15),
            ("Recommended Type", "recommended_type", 15),
            ("Current Monthly", "current_monthly", 15),
            ("Monthly Savings", "monthly_savings", 15),
            ("Confidence", "confidence", 12),
        ]

        # Headers
        for col, (header, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            ws.column_dimensions[cell.column_letter].width = width

        # Data rows
        for row_num, server in enumerate(servers, 2):
            for col, (_, field, _) in enumerate(columns, 1):
                value = server.get(field)
                cell = ws.cell(row=row_num, column=col, value=value)

                if field == "classification":
                    if value == "oversized":
                        cell.fill = self.OVERSIZED_FILL
                    elif value == "undersized":
                        cell.fill = self.UNDERSIZED_FILL

        ws.freeze_panes = "B2"


def generate_multi_account_report(
    summary: MultiAccountSummary,
    report_builder: MultiAccountReportBuilder,
    output_path: str
) -> Path:
    """Generate a multi-account Excel report.

    Args:
        summary: MultiAccountSummary from analysis
        report_builder: MultiAccountReportBuilder with all data
        output_path: Path for output file

    Returns:
        Path to generated file
    """
    generator = MultiAccountExcelGenerator(output_path)
    return generator.generate(report_builder)
